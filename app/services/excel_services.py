import re
import logging
from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook

log = logging.getLogger(__name__)

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

ABA_COMITENTES = "COMITENTES"
COL_CPF        = "E"
COL_STATUS     = "I"
COL_VOTOS_INI  = 12   # coluna L

ASSESSORES_LEGAIS = {
    "PGA":  "Pinheiro Guimarães Advogados & CSW Advogados",
    "MM":   "Machado Meyer Advogados",
    "FEL":  "Felsberg Advogados",
    "CTP":  "Costa Tavares Paes Advogados",
}

ASSESSORES_FINANCEIROS = {
    "BR":   "BR partners",
    "G5":   "G5 partners consultoria",
    "JNEY": "Journey Capital",
    "VIR":  "Virtus",
}

MAPA_VOTOS = {
    "A":  "sim",
    "R":  "não",
    "AB": "ab",
    "NV": "nv",
}

# =============================================================================
# TIPOS
# =============================================================================

@dataclass
class DadosArquivo:
    cpf: str          # 11 dígitos, sem formatação
    votos: list[str]
    arquivo: Path

@dataclass
class ExcelResult:
    preenchidos: int
    pulados: int      # já tinham status "ok"
    nao_encontrados: list[DadosArquivo]

# =============================================================================
# PARSING DO NOME DO ARQUIVO
# =============================================================================

def _normalizar_cpf(valor: str) -> str:
    return re.sub(r"\D", "", str(valor)).zfill(11)

def _parsear_nome(caminho: Path) -> DadosArquivo | None:
    stem = caminho.stem.strip().replace('"', "")
    match = re.match(
        r"^(\d{3}[\.\-\s]?\d{3}[\.\-\s]?\d{3}[\.\-\s]?\d{2})\s*-\s*(.+)$",
        stem,
    )
    if not match:
        log.warning("Nome fora do padrão: %s", caminho.name)
        return None

    cpf   = _normalizar_cpf(match.group(1))
    votos = [v.strip() for v in match.group(2).split(",") if v.strip()]
    return DadosArquivo(cpf=cpf, votos=votos, arquivo=caminho)

def carregar_arquivos(pasta: Path) -> dict[str, DadosArquivo]:
    """
    Lê os nomes dos PDFs já renomeados e monta um dicionário CPF → DadosArquivo.
    Não abre nenhum PDF — só lê o nome do arquivo.
    """
    mapa: dict[str, DadosArquivo] = {}
    for pdf in pasta.rglob("*.pdf"):
        dados = _parsear_nome(pdf)
        if dados:
            if dados.cpf in mapa:
                log.warning("CPF duplicado nos arquivos: %s", dados.cpf)
            mapa[dados.cpf] = dados
    return mapa

# =============================================================================
# TRADUÇÃO DE VOTOS
# =============================================================================

def _traduzir_voto(valor: str) -> str:
    v = re.sub(r"[^A-Z0-9]", "", valor.strip().upper())
    return (
        MAPA_VOTOS.get(v)
        or ASSESSORES_LEGAIS.get(v)
        or ASSESSORES_FINANCEIROS.get(v)
        or valor
    )

# =============================================================================
# EXCEL
# =============================================================================

def _detectar_num_colunas_votos(ws, col_inicio: int) -> int:
    col = col_inicio
    while True:
        valor = ws.cell(row=1, column=col).value
        if valor is None:
            break
        if isinstance(valor, str) and "SÉRIE" in valor.upper():
            break
        col += 1
    return col - col_inicio

def preencher_excel(excel_path: Path, mapa: dict[str, DadosArquivo]) -> ExcelResult:
    """
    Preenche a planilha Excel com os votos dos arquivos renomeados.
    Lookup O(1) por CPF — sem loop linear.
    """
    wb = load_workbook(excel_path)
    ws = wb[ABA_COMITENTES]

    num_colunas = _detectar_num_colunas_votos(ws, COL_VOTOS_INI)
    preenchidos = 0
    pulados     = 0
    usados: set[str] = set()

    linha = 2
    while True:
        cpf_raw = ws[f"{COL_CPF}{linha}"].value
        if cpf_raw is None:
            break

        cpf_norm = _normalizar_cpf(str(cpf_raw))

        # já processado anteriormente
        if ws[f"{COL_STATUS}{linha}"].value == "ok":
            usados.add(cpf_norm)
            pulados += 1
            linha += 1
            continue

        dados = mapa.get(cpf_norm)   # ← O(1), sem loop

        if dados:
            ws[f"{COL_STATUS}{linha}"] = "ok"
            for i, voto in enumerate(dados.votos[:num_colunas]):
                celula = ws.cell(row=linha, column=COL_VOTOS_INI + i)
                if celula.value is None:
                    celula.value = _traduzir_voto(voto)
            usados.add(cpf_norm)
            preenchidos += 1

        linha += 1

    wb.save(excel_path)

    nao_encontrados = [
        d for cpf, d in mapa.items() if cpf not in usados
    ]

    return ExcelResult(
        preenchidos=preenchidos,
        pulados=pulados,
        nao_encontrados=nao_encontrados,
    )